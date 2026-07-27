"""答题卡 Excel 报告导出路由（从 answer_sheet_grading_routes.py 拆分）

包含：
- 单份答题卡 Excel 报告导出（题目明细 + 错题汇总 + 归因报告 三个 Sheet）
- 批量答题卡 Excel 报告导出（ZIP 包）
"""
import os
import json
import logging

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from backend.core.database import get_db
from backend.core.security import get_current_user, assert_owner_or_admin
from backend.models.tables import RegisteredPerson, Exam

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/answer-sheet", tags=["answer-sheet"])

_EXCEL_HEADER_FILL = None


def _init_excel_styles():
    """延迟初始化 Excel 样式（避免 import 时失败）"""
    global _EXCEL_HEADER_FILL
    if _EXCEL_HEADER_FILL is not None:
        return
    from openpyxl.styles import PatternFill, Font, Alignment
    _EXCEL_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def _build_excel_workbook(db: Session, submission_id: int):
    """生成单个 submission 的 Excel Workbook（D 方案抽取的辅助函数）

    被 export_excel_report（单导出）和 export_excel_batch（批量 ZIP）共享复用。

    Args:
        db: 数据库会话
        submission_id: 提交 ID

    Returns:
        (workbook, student_name, exam_title, filename)

    Raises:
        HTTPException: 404 submission 不存在
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from backend.models.tables import ExamSubmission, Exam, Answer, RegisteredPerson, Question, KnowledgeAnalysis

    # 查询 submission
    submission = db.query(ExamSubmission).filter(ExamSubmission.id == submission_id).first()
    if not submission:
        raise HTTPException(404, f"提交 {submission_id} 不存在")

    # 查询考试和学生信息
    exam = db.query(Exam).filter(Exam.id == submission.exam_id).first()
    student = db.query(RegisteredPerson).filter(RegisteredPerson.id == submission.student_id).first()
    student_name = student.name if student else f"用户#{submission.student_id}"
    exam_title = exam.title if exam else f"考试#{submission.exam_id}"

    # 查询所有答案（按题目顺序）
    answers_query = (
        db.query(Answer, Question)
        .join(Question, Answer.question_id == Question.id)
        .filter(Answer.submission_id == submission_id)
        .order_by(Question.order)
        .all()
    )

    # 类型中文化
    type_map = {"single": "单选题", "multi": "多选题", "judge": "判断题", "fill": "填空题", "essay": "解答题/作文"}
    correct_map = {True: "✓ 正确", False: "✗ 错误", None: "— 未批改"}

    # 创建 Excel
    _init_excel_styles()
    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # === Sheet 1: 题目明细 ===
    ws1 = wb.active
    ws1.title = "题目明细"

    # 顶部信息
    ws1["A1"] = "学生姓名"
    ws1["B1"] = student_name
    ws1["C1"] = "考试名称"
    ws1["D1"] = exam_title
    ws1["E1"] = "总分"
    ws1["F1"] = f"{submission.score or 0}/{exam.total_score if exam else 100}"
    for col in "ABCDEF":
        ws1[f"{col}1"].font = Font(bold=True)

    # 表头（第3行）
    headers1 = ["题号", "题型", "题目内容", "学生答案", "标准答案", "得分", "满分", "批改状态", "错因", "评语"]
    for col_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 数据行
    for i, (ans, q) in enumerate(answers_query, start=4):
        # 学生答案：如果是选项索引，转为 A/B/C/D
        student_ans = ans.content or ""
        if q.type in ("single", "multi") and student_ans:
            try:
                if q.type == "single":
                    student_ans = chr(ord("A") + int(student_ans))
                else:
                    student_ans = "".join(chr(ord("A") + int(x)) for x in student_ans.split(","))
            except (ValueError, IndexError):
                pass  # 保留原始内容
        elif q.type == "judge" and student_ans:
            student_ans = "正确" if student_ans.lower() == "true" else "错误"

        # 标准答案同样转换
        std_ans = q.answer or ""
        if q.type in ("single", "multi") and std_ans:
            try:
                if q.type == "single":
                    std_ans = chr(ord("A") + int(std_ans))
                else:
                    std_ans = "".join(chr(ord("A") + int(x)) for x in std_ans.split(","))
            except (ValueError, IndexError):
                pass
        elif q.type == "judge" and std_ans:
            std_ans = "正确" if std_ans.lower() == "true" else "错误"

        row_data = [
            q.order,
            type_map.get(q.type, q.type),
            q.content[:200],
            student_ans,
            std_ans,
            ans.score or 0,
            q.score,
            correct_map.get(ans.is_correct, "— 未批改"),
            "",  # 错因（Answer 表没有，留空）
            "",  # 评语（Answer 表没有，留空）
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=i, column=col_idx, value=val)
            cell.alignment = left_align if col_idx in (3, 4, 5, 9, 10) else center_align
            if ans.is_correct is False and col_idx == 8:
                cell.fill = wrong_fill

    # 列宽
    col_widths = [6, 10, 50, 20, 20, 8, 8, 12, 15, 40]
    for idx, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = w

    # === Sheet 2: 错题汇总 ===
    ws2 = wb.create_sheet("错题汇总")
    ws2["A1"] = f"{student_name} 的错题汇总"
    ws2["A1"].font = Font(bold=True, size=14)
    headers2 = ["题号", "题型", "题目内容", "学生答案", "标准答案", "错因", "改进建议"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    row_idx = 4
    for ans, q in answers_query:
        if ans.is_correct is not False:
            continue
        # 学生答案转换
        student_ans = ans.content or ""
        if q.type in ("single", "multi") and student_ans:
            try:
                if q.type == "single":
                    student_ans = chr(ord("A") + int(student_ans))
                else:
                    student_ans = "".join(chr(ord("A") + int(x)) for x in student_ans.split(","))
            except (ValueError, IndexError):
                pass
        elif q.type == "judge" and student_ans:
            student_ans = "正确" if student_ans.lower() == "true" else "错误"
        std_ans = q.answer or ""

        row_data = [q.order, type_map.get(q.type, q.type), q.content[:200], student_ans, std_ans, "（详见归因报告）", "（详见归因报告）"]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = left_align if col_idx in (3, 4, 5, 6, 7) else center_align
        row_idx += 1

    if row_idx == 4:
        ws2.cell(row=4, column=1, value="（无错题）")

    col_widths2 = [6, 10, 50, 20, 20, 20, 40]
    for idx, w in enumerate(col_widths2, start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # === Sheet 3: 归因报告 ===
    ws3 = wb.create_sheet("归因报告")
    ws3["A1"] = f"{student_name} 的学情归因报告"
    ws3["A1"].font = Font(bold=True, size=14)

    analyses = (
        db.query(KnowledgeAnalysis)
        .filter(KnowledgeAnalysis.student_id == submission.student_id)
        .order_by(KnowledgeAnalysis.created_at.desc())
        .all()
    )

    if not analyses:
        ws3.cell(row=3, column=1, value="暂无归因分析数据")
    else:
        row_idx = 3
        for ka in analyses:
            # 分析类型标题
            type_label = "数学学情" if ka.analysis_type == "math" else "写作学情"
            ws3.cell(row=row_idx, column=1, value=f"【{type_label}】（生成时间：{ka.created_at.strftime('%Y-%m-%d %H:%M')}）")
            ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            row_idx += 1

            # 雷达图数据
            try:
                radar = json.loads(ka.radar_json) if ka.radar_json else {}
            except json.JSONDecodeError:
                radar = {}
            if radar:
                ws3.cell(row=row_idx, column=1, value="维度掌握度：")
                ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                row_idx += 1
                for dim_name, score in radar.items():
                    ws3.cell(row=row_idx, column=1, value=dim_name)
                    ws3.cell(row=row_idx, column=2, value=f"{score*100:.0f}%")
                    row_idx += 1
                row_idx += 1

            # 薄弱点
            try:
                weak_points = json.loads(ka.weak_points_json) if ka.weak_points_json else {}
            except json.JSONDecodeError:
                weak_points = {}

            if ka.analysis_type == "math":
                # 数学：weak_points 是列表
                if isinstance(weak_points, list) and weak_points:
                    ws3.cell(row=row_idx, column=1, value="薄弱知识点：")
                    ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                    row_idx += 1
                    # 表头
                    for col_idx, h in enumerate(["知识点", "薄弱度", "错题数", "改进建议"], start=1):
                        cell = ws3.cell(row=row_idx, column=col_idx, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                    row_idx += 1
                    for wp in weak_points:
                        ws3.cell(row=row_idx, column=1, value=wp.get("knowledge_name", ""))
                        ws3.cell(row=row_idx, column=2, value=f"{wp.get('weakness_score', 0)*100:.0f}%")
                        ws3.cell(row=row_idx, column=3, value=wp.get("error_count", 0))
                        ws3.cell(row=row_idx, column=4, value=wp.get("suggestion", ""))
                        row_idx += 1
            else:
                # 写作：weak_points 是 dict（含 weak_dimensions/error_cause_distribution/overall_suggestion）
                if isinstance(weak_points, dict):
                    weak_dims = weak_points.get("weak_dimensions", [])
                    if weak_dims:
                        ws3.cell(row=row_idx, column=1, value="薄弱维度：")
                        ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                        row_idx += 1
                        for col_idx, h in enumerate(["维度", "薄弱度", "错因", "改进建议"], start=1):
                            cell = ws3.cell(row=row_idx, column=col_idx, value=h)
                            cell.fill = header_fill
                            cell.font = header_font
                        row_idx += 1
                        for wd in weak_dims:
                            ws3.cell(row=row_idx, column=1, value=wd.get("dimension_name", ""))
                            ws3.cell(row=row_idx, column=2, value=f"{wd.get('weakness_score', 0)*100:.0f}%")
                            ws3.cell(row=row_idx, column=3, value=", ".join(wd.get("error_causes", [])))
                            ws3.cell(row=row_idx, column=4, value=wd.get("suggestion", ""))
                            row_idx += 1
                    overall = weak_points.get("overall_suggestion", "")
                    if overall:
                        row_idx += 1
                        ws3.cell(row=row_idx, column=1, value="综合建议：")
                        ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                        row_idx += 1
                        ws3.cell(row=row_idx, column=1, value=overall)
                        ws3.cell(row=row_idx, column=1).alignment = left_align
                        ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                        row_idx += 1

            row_idx += 2  # 不同分析类型之间空一行

    col_widths3 = [25, 15, 25, 50]
    for idx, w in enumerate(col_widths3, start=1):
        ws3.column_dimensions[get_column_letter(idx)].width = w

    # 文件名（中文安全）
    safe_student = student_name.replace(" ", "_").replace("/", "_")
    filename = f"答题卡报告_{safe_student}_{submission_id}.xlsx"

    return wb, student_name, exam_title, filename


@router.get("/export/excel/{submission_id}")
def export_excel_report(
    submission_id: int,
    current_user: RegisteredPerson = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """将整卷扫描批改结果导出为 Excel 报告

    包含 3 个 Sheet：
    1. 题目明细：题号、题型、题目内容、学生答案、标准答案、得分、是否正确、错因、评语
    2. 错题汇总：错题题号、题型、错因、改进建议
    3. 归因报告：薄弱知识点/写作薄弱维度 + 雷达图数据
    """
    import io
    from fastapi.responses import StreamingResponse

    if current_user.role not in ("teacher", "admin"):
        raise HTTPException(403, "仅教师/管理员可导出报告")

    # 校验提交归属（通过 submission → exam → teacher）
    from backend.models.tables import ExamSubmission
    submission = db.query(ExamSubmission).filter(ExamSubmission.id == submission_id).first()
    if not submission:
        raise HTTPException(404, f"提交 {submission_id} 不存在")
    exam = db.query(Exam).filter(Exam.id == submission.exam_id).first()
    if exam:
        assert_owner_or_admin(exam.teacher_id, current_user)

    wb, student_name, exam_title, filename = _build_excel_workbook(db, submission_id)

    # 保存到内存
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/export/excel-batch")
def export_excel_batch(
    submission_ids: str,
    current_user: RegisteredPerson = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """D. 批量 Excel 报告导出（ZIP 包）

    一次导出多个 submission 的 Excel 报告，打包为 ZIP。

    Args:
        submission_ids: 逗号分隔的 submission ID 列表（最多 100 个）

    Returns:
        application/zip 流，包含 N 个 .xlsx 文件
    """
    import io
    import zipfile
    from fastapi.responses import StreamingResponse
    from backend.models.tables import ExamSubmission

    if current_user.role not in ("teacher", "admin"):
        raise HTTPException(403, "仅教师/管理员可批量导出报告")

    # 解析 submission_ids
    try:
        sid_list = [int(s.strip()) for s in submission_ids.split(",") if s.strip()]
    except ValueError:
        raise HTTPException(400, "submission_ids 格式错误，需为逗号分隔的整数列表")

    if not sid_list:
        raise HTTPException(400, "submission_ids 不能为空")
    if len(sid_list) > 100:
        raise HTTPException(400, f"单批最多导出 100 份报告，当前 {len(sid_list)} 份")

    # 校验所有 submission 存在 + 归属当前教师
    submissions = db.query(ExamSubmission).filter(ExamSubmission.id.in_(sid_list)).all()
    sub_map = {s.id: s for s in submissions}
    missing = [sid for sid in sid_list if sid not in sub_map]
    if missing:
        raise HTTPException(404, f"以下 submission 不存在: {missing[:10]}{'...' if len(missing) > 10 else ''}")

    # 校验每个 submission 的考试归属
    exam_ids = list(set(s.exam_id for s in submissions))
    exam_map = {e.id: e for e in db.query(Exam).filter(Exam.id.in_(exam_ids)).all()} if exam_ids else {}
    for sub in submissions:
        exam = exam_map.get(sub.exam_id)
        if exam:
            assert_owner_or_admin(exam.teacher_id, current_user)

    # 逐个生成 Excel 并打包成 ZIP
    zip_buf = io.BytesIO()
    used_filenames: set[str] = set()  # 处理同名冲突
    success_count = 0
    failed_list: list[dict] = []

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for sid in sid_list:
            try:
                wb, student_name, exam_title, original_filename = _build_excel_workbook(db, sid)
                # 处理同名冲突：如有重名，附加 _2/_3
                filename = original_filename
                if filename in used_filenames:
                    base, ext = os.path.splitext(original_filename)
                    n = 2
                    while f"{base}_{n}{ext}" in used_filenames:
                        n += 1
                    filename = f"{base}_{n}{ext}"
                used_filenames.add(filename)

                excel_buf = io.BytesIO()
                wb.save(excel_buf)
                excel_buf.seek(0)
                zf.writestr(filename, excel_buf.read())
                success_count += 1
            except HTTPException as e:
                failed_list.append({"submission_id": sid, "error": e.detail, "status_code": e.status_code})
            except Exception as e:
                failed_list.append({"submission_id": sid, "error": f"{type(e).__name__}: {e}"})
                logger.exception(
                    f"[AnswerSheetRoute] 批量导出 submission={sid} 异常: {e}"
                )

    zip_buf.seek(0)

    # 如果全部失败，返回错误
    if success_count == 0:
        raise HTTPException(
            500,
            f"所有 {len(sid_list)} 份报告生成失败: {failed_list[:3]}"
        )

    # ZIP 文件名（含时间戳）
    from datetime import datetime as _dt
    zip_filename = f"答题卡报告_batch_{_dt.now().strftime('%Y%m%d_%H%M%S')}.zip"

    logger.info(
        f"[AnswerSheetRoute] 批量导出完成: total={len(sid_list)}, "
        f"success={success_count}, failed={len(failed_list)}"
    )

    # 注意：失败的 submission 信息记录在 response header 中（避免破坏 ZIP 流）
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{zip_filename}",
        "X-Batch-Total": str(len(sid_list)),
        "X-Batch-Success": str(success_count),
        "X-Batch-Failed": str(len(failed_list)),
    }
    if failed_list:
        # 把失败列表序列化到 header（截断防止 header 过大）
        failed_summary = json.dumps(failed_list[:5], ensure_ascii=False)[:500]
        headers["X-Batch-Failed-Detail"] = failed_summary

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers=headers,
    )
