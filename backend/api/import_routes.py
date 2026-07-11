"""批量导入 API — Excel 模板下载 + 批量导入 + 部门管理"""

import io
import re
from typing import List

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body
from sqlalchemy.orm import Session

from backend.core.database import get_db
from backend.core.security import get_current_user, require_role
from backend.core.config import settings
from backend.models.tables import RegisteredPerson, Department
from backend.models.schemas import (
    DepartmentOut, ImportResult, ImportResultRow, PersonOut,
)
from backend.core.security import hash_password

router = APIRouter(prefix="/api/import", tags=["import"])

PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
EMAIL_RE = re.compile(r"^[\w.-]+@[\w.-]+\.\w+$")
MAX_ROWS = 1000
DEFAULT_PASSWORD = "123456"

# ===== 部门管理 =====

@router.get("/departments", response_model=List[DepartmentOut])
def list_departments(
    current_user: RegisteredPerson = Depends(require_role(["admin"])),
    db: Session = Depends(get_db),
):
    deps = db.query(Department).order_by(Department.name).all()
    result = []
    for d in deps:
        count = db.query(RegisteredPerson).filter(RegisteredPerson.department_id == d.id).count()
        result.append(DepartmentOut(
            id=d.id, name=d.name, type=d.type,
            member_count=count, created_at=d.created_at,
        ))
    return result


@router.post("/departments", response_model=DepartmentOut)
def create_department(
    name: str,
    type: str = "class",
    current_user: RegisteredPerson = Depends(require_role(["admin"])),
    db: Session = Depends(get_db),
):
    existing = db.query(Department).filter(Department.name == name).first()
    if existing:
        raise HTTPException(400, f"部门/班级 '{name}' 已存在")
    dep = Department(name=name, type=type)
    db.add(dep)
    db.commit()
    db.refresh(dep)
    return DepartmentOut(id=dep.id, name=dep.name, type=dep.type, member_count=0, created_at=dep.created_at)


@router.delete("/departments/{department_id}")
def delete_department(
    department_id: int,
    current_user: RegisteredPerson = Depends(require_role(["admin"])),
    db: Session = Depends(get_db),
):
    dep = db.query(Department).filter(Department.id == department_id).first()
    if not dep:
        raise HTTPException(404, "部门不存在")
    members = db.query(RegisteredPerson).filter(RegisteredPerson.department_id == department_id).count()
    if members > 0:
        raise HTTPException(400, f"该部门下有 {members} 名成员，无法删除")
    db.delete(dep)
    db.commit()
    return {"message": "部门已删除"}


# ===== Excel 模板下载 =====

@router.get("/template")
def download_template(
    role: str = "student",
    current_user: RegisteredPerson = Depends(require_role(["admin"])),
):
    """下载标准 Excel 导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active

    if role == "student":
        ws.title = "学生花名册"
        headers = ["学号*", "姓名*", "手机号*", "班级", "专业", "邮箱", "身份证"]
        example = ["2024001", "张三", "13800138000", "计算机1班", "软件工程", "zhangsan@edu.cn", "110101200001011234"]
    elif role == "teacher":
        ws.title = "教师花名册"
        headers = ["工号*", "姓名*", "手机号*", "部门", "专业", "邮箱", "身份证"]
        example = ["T001", "李老师", "13900139000", "计算机系", "人工智能", "liteacher@edu.cn", ""]
    else:
        ws.title = "管理员花名册"
        headers = ["工号*", "姓名*", "手机号*", "部门", "邮箱"]
        example = ["A001", "王管理", "13700137000", "教务处", "admin@edu.cn"]

    ws.append(headers)
    ws.append(example)

    # 设置列宽
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(h) * 2, 15)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=import_template_{role}.xlsx"},
    )


# ===== 批量导入核心逻辑 =====

@router.post("/excel", response_model=ImportResult)
async def import_excel(
    file: UploadFile = File(...),
    role: str = "student",
    current_user: RegisteredPerson = Depends(require_role(["admin"])),
    db: Session = Depends(get_db),
):
    """批量导入 Excel 文件"""
    if role not in ("student", "teacher", "admin"):
        raise HTTPException(400, "无效角色")

    # 文件校验
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx/.xls 文件")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "文件大小不能超过 10MB")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    wb.close()

    if len(rows) > MAX_ROWS:
        raise HTTPException(400, f"单次最多导入 {MAX_ROWS} 行，当前 {len(rows)} 行")

    return _process_import(rows, role, db)


@router.post("/csv", response_model=ImportResult)
async def import_csv(
    data: str = Body(..., media_type="text/plain"),
    role: str = "student",
    current_user: RegisteredPerson = Depends(require_role(["admin"])),
    db: Session = Depends(get_db),
):
    """复制粘贴批量导入（CSV 格式文本）"""
    if role not in ("student", "teacher", "admin"):
        raise HTTPException(400, "无效角色")

    lines = data.strip().split("\n")
    rows = []
    for line in lines:
        if not line.strip():
            continue
        # 支持逗号和制表符分隔
        if "\t" in line:
            cells = line.split("\t")
        else:
            cells = line.split(",")
        rows.append([c.strip() for c in cells])

    if len(rows) > MAX_ROWS:
        raise HTTPException(400, f"单次最多导入 {MAX_ROWS} 行，当前 {len(rows)} 行")

    return _process_import(rows, role, db)


def _process_import(rows: list, role: str, db: Session) -> ImportResult:
    """核心导入处理：校验 + 分批入库"""
    errors: list[ImportResultRow] = []
    success_count = 0

    # 预加载已有数据用于唯一性校验
    existing_employee_ids = {r.employee_id for r in db.query(RegisteredPerson.employee_id).all() if r.employee_id}
    existing_usernames = {r.username for r in db.query(RegisteredPerson.username).all() if r.username}
    existing_phones = {r.phone for r in db.query(RegisteredPerson.phone).all() if r.phone}
    departments_cache = {d.name: d.id for d in db.query(Department).all()}

    # 本批次内去重
    batch_employee_ids = set()
    batch_usernames = set()

    valid_records = []

    for idx, row in enumerate(rows, start=2):  # Excel 行号从2开始（1是表头）
        # 补齐空列
        while len(row) < 7:
            row = list(row) + [None]

        employee_id = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        phone = str(row[2] or "").strip()
        dept_name = str(row[3] or "").strip()
        major = str(row[4] or "").strip() if len(row) > 4 else ""
        email = str(row[5] or "").strip() if len(row) > 5 else ""
        id_card = str(row[6] or "").strip() if len(row) > 6 else ""

        row_errors = []

        # 必填校验
        if not employee_id:
            row_errors.append("学号/工号不能为空")
        if not name:
            row_errors.append("姓名不能为空")
        if not phone:
            row_errors.append("手机号不能为空")

        # 格式校验
        if phone and not PHONE_RE.match(phone):
            row_errors.append("手机号格式错误（需11位国内手机号）")
        if email and not EMAIL_RE.match(email):
            row_errors.append("邮箱格式错误")

        # 唯一性校验
        if employee_id:
            if employee_id in existing_employee_ids or employee_id in batch_employee_ids:
                row_errors.append(f"学号/工号 '{employee_id}' 已存在")
            else:
                batch_employee_ids.add(employee_id)

        if phone and phone in existing_phones:
            row_errors.append(f"手机号 '{phone}' 已被注册")

        # 用户名 = 学号/工号（自动生成）
        username = employee_id
        if username and (username in existing_usernames or username in batch_usernames):
            row_errors.append(f"用户名 '{username}' 已存在")
        elif username:
            batch_usernames.add(username)

        if row_errors:
            errors.append(ImportResultRow(
                row=idx, employee_id=employee_id, name=name,
                error="; ".join(row_errors),
            ))
            continue

        # 解析部门
        department_id = None
        if dept_name:
            if dept_name in departments_cache:
                department_id = departments_cache[dept_name]
            else:
                # 自动创建部门
                dep = Department(name=dept_name, type="class" if role == "student" else "department")
                db.add(dep)
                db.flush()
                departments_cache[dept_name] = dep.id
                department_id = dep.id

        valid_records.append({
            "employee_id": employee_id,
            "name": name,
            "role": role,
            "username": username,
            "phone": phone,
            "department_id": department_id,
            "major": major,
            "email": email,
            "id_card": id_card,
        })

    # 分批入库
    BATCH_SIZE = 200
    for i in range(0, len(valid_records), BATCH_SIZE):
        batch = valid_records[i:i + BATCH_SIZE]
        for rec in batch:
            person = RegisteredPerson(
                name=rec["name"],
                role=rec["role"],
                username=rec["username"],
                password_hash=hash_password(DEFAULT_PASSWORD),
                face_embedding="[]",  # 无人脸数据
                employee_id=rec["employee_id"],
                phone=rec["phone"],
                department_id=rec["department_id"],
                major=rec["major"],
                email=rec["email"],
                id_card=rec["id_card"],
            )
            db.add(person)
        db.commit()
        success_count += len(batch)

    return ImportResult(
        total=len(rows),
        success=success_count,
        failed=len(errors),
        errors=errors,
    )


# ===== 导出错误明细 =====

@router.post("/export-errors")
def export_errors(
    errors: List[ImportResultRow],
    current_user: RegisteredPerson = Depends(require_role(["admin"])),
):
    """将导入错误明细导出为 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导入错误明细"
    ws.append(["行号", "学号/工号", "姓名", "错误原因"])

    for e in errors:
        ws.append([e.row, e.employee_id, e.name, e.error])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_errors.xlsx"},
    )
