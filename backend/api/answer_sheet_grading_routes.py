"""答题卡批改与导出路由（从 answer_sheet_routes.py 拆分）"""
import os
import json
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from backend.core.database import get_db
from backend.core.security import get_current_user, assert_owner_or_admin
from backend.models.tables import RegisteredPerson, Exam, Question
from backend.services.answer_sheet import answer_sheet_orchestrator
from backend.services.paper_template import paper_template_service
from cv_engine.detectors.answer_card_detector import answer_card_detector

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/answer-sheet", tags=["answer-sheet"])

SCAN_UPLOAD_DIR = "uploads/answer_sheets"
os.makedirs(SCAN_UPLOAD_DIR, exist_ok=True)


@router.post("/submissions/{submission_id}/questions/{question_id}/manual-input")
def manual_input_answer(
    submission_id: int,
    question_id: int,
    student_answer: str = Form(...),
    current_user: RegisteredPerson = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """教师人工补录学生答案并重新判分（D 方案）

    场景：当 OCR 双引擎均失败 needs_manual_input=True、或 OCR 置信度过低
    导致 is_correct=None 时，教师查看扫描件后手动输入学生答案，
    系统重新调 auto_grade 判分并更新 Answer 表 + Submission 总分。

    权限：仅教师/管理员（且必须是该 submission 所属考试的教师）

    流程：
    1. 校验 submission 存在 + 当前用户是该考试的教师
    2. 校验 question 存在 + 属于该考试
    3. 调 auto_grade(question, student_answer) 重新判分
       - 填空题：会走 A+B 方案（多空拆分 + 数值/单位容差）
       - 选择/判断题：常规精确匹配
       - 大题：返回 (0, False)，大题应走 LLM 重批改接口
    4. 更新或新建 Answer 记录（content/score/is_correct）
    5. 重新计算 submission 总分（所有 answers 的 score 之和）
    6. 更新 submission.graded_at
    7. 返回更新后的批改结果
    """
    from datetime import datetime
    from backend.models.tables import ExamSubmission, Answer

    if current_user.role not in ("teacher", "admin"):
        raise HTTPException(403, "仅教师/管理员可调用此接口")

    # 1. 校验 submission
    submission = db.query(ExamSubmission).filter(ExamSubmission.id == submission_id).first()
    if not submission:
        raise HTTPException(404, f"提交 {submission_id} 不存在")

    # 2. 校验当前用户是该 submission 所属考试的教师
    exam = db.query(Exam).filter(Exam.id == submission.exam_id).first()
    if not exam:
        raise HTTPException(404, "提交关联的考试不存在")
    if exam.teacher_id != current_user.id and current_user.role != "admin":
        raise HTTPException(403, "无权操作此提交（仅该考试的教师可操作）")

    # 3. 校验 question 存在 + 属于该考试
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(404, f"题目 {question_id} 不存在")
    if question.exam_id != submission.exam_id:
        raise HTTPException(400, "题目不属于该提交对应的考试")

    # 4. 大题不支持人工补录（应走 LLM 重批改）
    if question.type == "essay":
        raise HTTPException(
            400,
            "大题/作文不支持人工补录答案，请使用 LLM 重批改接口"
        )

    # 5. 规范化输入
    student_answer = (student_answer or "").strip()
    if not student_answer:
        raise HTTPException(400, "学生答案不能为空")

    # 6. 调 auto_grade 重新判分（填空题会走 A+B 方案）
    from backend.services.exam_service import auto_grade
    score, is_correct = auto_grade(question, student_answer)

    # 7. 更新或新建 Answer（先捕获 before 值用于历史记录）
    answer = db.query(Answer).filter(
        Answer.submission_id == submission_id,
        Answer.question_id == question_id,
    ).first()

    # F 方案：捕获重批改前的状态用于历史记录
    before_score = answer.score if answer else None
    before_is_correct = answer.is_correct if answer else None
    before_total_score = submission.score

    if answer:
        answer.content = student_answer
        answer.score = score
        answer.is_correct = is_correct
    else:
        answer = Answer(
            submission_id=submission_id,
            question_id=question_id,
            content=student_answer,
            score=score,
            is_correct=is_correct,
        )
        db.add(answer)

    # 8. 重新计算 submission 总分
    all_answers = db.query(Answer).filter(Answer.submission_id == submission_id).all()
    total_score = sum((a.score or 0) for a in all_answers)
    submission.score = total_score
    submission.graded_at = datetime.now()
    if submission.status != "graded":
        submission.status = "graded"

    # 9. F 方案：写入重批改历史记录（与 Answer 同事务提交）
    from backend.models.tables import AnswerRegradeHistory
    history = AnswerRegradeHistory(
        submission_id=submission_id,
        question_id=question_id,
        operator_id=current_user.id,
        regrade_method="manual_input",
        input_mode=None,
        force_essay=False,
        before_score=before_score,
        after_score=float(score),
        before_is_correct=before_is_correct,
        after_is_correct=is_correct,
        max_score=float(question.score),
        before_total_score=before_total_score,
        after_total_score=float(total_score),
        student_text=student_answer,
        is_essay=False,
        model_key=None,
        grading_method=None,
        error_cause=None,
        knowledge_points_json=None,
        grading_json=None,
        writing_attribution_json=None,
        comment=None,
    )
    db.add(history)

    db.commit()

    logger.info(
        f"[ManualInput] teacher={current_user.id} submission={submission_id} "
        f"question={question_id} answer={student_answer!r} → score={score}/{question.score}"
    )

    return {
        "submission_id": submission_id,
        "question_id": question_id,
        "student_answer": student_answer,
        "standard_answer": question.answer,
        "score": score,
        "max_score": question.score,
        "is_correct": is_correct,
        "total_score": total_score,
        "manual_input": True,
        "graded_at": submission.graded_at.isoformat() if submission.graded_at else None,
    }



@router.post("/submissions/{submission_id}/questions/{question_id}/regrade-essay")
async def regrade_essay(
    submission_id: int,
    question_id: int,
    student_text: Optional[str] = Form(None),
    image_file: Optional[UploadFile] = File(None),
    force_essay: bool = Form(False),
    current_user: RegisteredPerson = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """大题/作文 LLM 重批改接口

    场景：scan_and_grade 完成后，教师对某道大题想重新触发 LLM 批改：
    - 上次 OCR 误识别 → 教师直接输入学生答案文字（student_text），跳过 OCR
    - 上次 LLM 评分不准 → 重新上传图片（image_file），重新走 OCR + LLM
    - 想换路由判定 → force_essay=True 强制按作文批改（默认按 _is_essay_question 自动路由）

    入参（student_text 和 image_file 二选一，至少一个）：
    - student_text: 教师直接输入的学生答案文字（跳过 OCR，confidence=1.0）
    - image_file: 重新上传的图片，重新走 OCR + LLM
    - force_essay: True 强制按作文批改；False（默认）按 _is_essay_question 自动路由

    权限：仅教师/管理员（且必须是该 submission 所属考试的教师）

    流程：
    1. 权限校验 + 参数校验（student_text/image_file 至少一个）
    2. 校验 submission/question 存在 + 题目属于该考试 + 必须是大题（type==essay）
    3. 准备学生答案文字：
       - student_text 模式：直接用，confidence=1.0
       - image_file 模式：调 ocr_service.recognize，OCR 失败返回 400 提示改用 student_text
    4. 路由：force_essay 或 _is_essay_question → grade_essay；否则 grade_math
    5. 调 grading_service.grade_essay/grade_math（复用 scan_and_grade 中的 LLM 链路）
    6. 作文场景调 writing_kg 做错因→维度/节点/建议归因
    7. 更新或新建 Answer 表（content/score/is_correct）
    8. 重新计算 submission 总分 + status='graded'
    9. 返回完整批改详情（含 grading/error_cause/knowledge_points/writing_attribution）
    """
    from datetime import datetime
    from backend.models.tables import ExamSubmission, Answer
    from backend.services.grader import grading_service
    from backend.services.answer_sheet import _is_essay_question
    from backend.services.writing_graph import writing_kg

    # 0. 权限校验
    if current_user.role not in ("teacher", "admin"):
        raise HTTPException(403, "仅教师/管理员可调用此接口")

    # 1. 参数校验：student_text 和 image_file 至少一个
    # 注意：FastAPI File(None) 默认值在直接调用时是 File 对象（非 None），
    # 用 hasattr(image_file, 'read') 区分真实 UploadFile 与默认值占位
    has_text = bool(student_text and student_text.strip())
    has_image = image_file is not None and hasattr(image_file, 'read')
    if not has_text and not has_image:
        raise HTTPException(400, "必须提供 student_text 或 image_file（二选一）")

    # 2. 校验 submission
    submission = db.query(ExamSubmission).filter(ExamSubmission.id == submission_id).first()
    if not submission:
        raise HTTPException(404, f"提交 {submission_id} 不存在")

    # 3. 校验当前用户是该 submission 所属考试的教师
    exam = db.query(Exam).filter(Exam.id == submission.exam_id).first()
    if not exam:
        raise HTTPException(404, "提交关联的考试不存在")
    if exam.teacher_id != current_user.id and current_user.role != "admin":
        raise HTTPException(403, "无权操作此提交（仅该考试的教师可操作）")

    # 4. 校验 question
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(404, f"题目 {question_id} 不存在")
    if question.exam_id != submission.exam_id:
        raise HTTPException(400, "题目不属于该提交对应的考试")

    # 5. 必须是大题（type==essay）
    if question.type != "essay":
        raise HTTPException(
            400,
            "该接口仅支持大题/作文重批改，其他题型请使用 manual-input 接口"
        )

    # 6. 准备学生答案文字 + 图片字节
    image_bytes = None
    if has_image:
        image_bytes = await image_file.read()
        if not image_bytes:
            raise HTTPException(400, "图片内容为空")

    if has_text:
        raw_text = student_text.strip()
        confidence = 1.0  # 教师手输，置信度满分
    else:
        # image_file 模式：走 OCR
        from backend.services.ocr import ocr_service
        try:
            ocr_result = await ocr_service.recognize(image_bytes)
        except Exception as e:
            logger.exception(f"[RegradeEssay] OCR 异常: {e}")
            raise HTTPException(500, f"OCR 调用失败: {type(e).__name__}: {e}")
        raw_text = ocr_result.text or ""
        confidence = float(ocr_result.confidence or 0.0)
        if ocr_result.needs_manual_input or (not raw_text.strip() and confidence < 0.4):
            raise HTTPException(
                400,
                f"OCR 识别失败或置信度过低 (confidence={confidence:.2f})，"
                f"请改用 student_text 直接输入学生答案"
            )

    # 7. 路由：作文 vs 数学
    # 注意：FastAPI Form(False) 默认值在直接调用时是 Form 对象（truthy），
    # 用 `force_essay is True` 严格判断，避免测试/直接调用场景下误路由
    is_essay = (force_essay is True) or _is_essay_question(question.content)
    logger.info(
        f"[RegradeEssay] submission={submission_id} question={question_id} "
        f"is_essay={is_essay} force_essay={force_essay!r} "
        f"content_head={question.content[:30]!r}"
    )

    # 8. 调 LLM 批改（复用 grading_service）
    try:
        if is_essay:
            llm_result = await grading_service.grade_essay(
                question=question.content,
                standard_answer=question.answer or "",
                student_answer_ocr=raw_text,
                total_score=int(question.score),
                confidence=confidence,
                image_bytes=image_bytes,
            )
        else:
            llm_result = await grading_service.grade_math(
                question=question.content,
                standard_answer=question.answer or "",
                student_answer_ocr=raw_text,
                total_score=int(question.score),
                confidence=confidence,
                image_bytes=image_bytes,
            )
    except Exception as e:
        logger.exception(f"[RegradeEssay] LLM 批改异常: {e}")
        raise HTTPException(500, f"LLM 批改失败: {type(e).__name__}: {e}")

    # 9. 提取结果
    suggested_score = float(llm_result.get("suggested_score", 0) or 0)
    max_score = float(llm_result.get("max_score", question.score) or question.score)
    grading = llm_result.get("grading", {}) or {}
    comment = llm_result.get("comment", "") or ""
    error_cause = grading.get("error_cause", "none")
    knowledge_points = grading.get("knowledge_points", []) or []
    model_key = llm_result.get("model_key", "standard")
    grading_method = grading.get("grading_method", "llm")
    ratio = suggested_score / max_score if max_score > 0 else 0
    is_correct = ratio >= 0.8

    # 10. 写作能力归因（仅作文场景，且有错因）
    writing_attribution = None
    if is_essay and error_cause and error_cause != "none":
        try:
            dimension = writing_kg.map_error_cause_to_dimension(error_cause)
            fine_nodes = writing_kg.map_error_cause_to_nodes(error_cause)
            suggestion = writing_kg.get_error_cause_suggestion(error_cause)
            writing_attribution = {
                "error_cause": error_cause,
                "dimension": dimension,
                "fine_nodes": fine_nodes,
                "knowledge_points": knowledge_points,
                "suggestion": suggestion,
            }
            if suggestion:
                comment = f"{comment}\n\n【改进建议】{suggestion}"
        except Exception as e:
            logger.warning(f"[RegradeEssay] 写作归因失败: {type(e).__name__}: {e}")

    # 11. 更新或新建 Answer（先捕获 before 值用于历史记录）
    answer = db.query(Answer).filter(
        Answer.submission_id == submission_id,
        Answer.question_id == question_id,
    ).first()

    # F 方案：捕获重批改前的状态用于历史记录
    before_score = answer.score if answer else None
    before_is_correct = answer.is_correct if answer else None
    before_total_score = submission.score

    if answer:
        answer.content = raw_text
        answer.score = suggested_score
        answer.is_correct = is_correct
    else:
        answer = Answer(
            submission_id=submission_id,
            question_id=question_id,
            content=raw_text,
            score=suggested_score,
            is_correct=is_correct,
        )
        db.add(answer)

    # 12. 重新计算 submission 总分
    all_answers = db.query(Answer).filter(Answer.submission_id == submission_id).all()
    total_score = sum((a.score or 0) for a in all_answers)
    submission.score = total_score
    submission.graded_at = datetime.now()
    if submission.status != "graded":
        submission.status = "graded"

    # 13. F 方案：写入重批改历史记录（与 Answer 同事务提交，保证审计一致性）
    from backend.models.tables import AnswerRegradeHistory
    history = AnswerRegradeHistory(
        submission_id=submission_id,
        question_id=question_id,
        operator_id=current_user.id,
        regrade_method="regrade_essay",
        input_mode="text" if has_text else "image",
        force_essay=(force_essay is True),  # 严格判断，避开 Form(False) 默认值陷阱
        before_score=before_score,
        after_score=suggested_score,
        before_is_correct=before_is_correct,
        after_is_correct=is_correct,
        max_score=max_score,
        before_total_score=before_total_score,
        after_total_score=total_score,
        student_text=raw_text,
        is_essay=is_essay,
        model_key=model_key,
        grading_method=grading_method,
        error_cause=error_cause,
        knowledge_points_json=json.dumps(knowledge_points, ensure_ascii=False) if knowledge_points else None,
        grading_json=json.dumps(llm_result, ensure_ascii=False),
        writing_attribution_json=json.dumps(writing_attribution, ensure_ascii=False) if writing_attribution else None,
        comment=comment,
    )
    db.add(history)

    db.commit()

    logger.info(
        f"[RegradeEssay] teacher={current_user.id} submission={submission_id} "
        f"question={question_id} is_essay={is_essay} → score={suggested_score}/{max_score} "
        f"model={model_key} error_cause={error_cause}"
    )

    return {
        "submission_id": submission_id,
        "question_id": question_id,
        "student_answer": raw_text,
        "standard_answer": question.answer,
        "score": suggested_score,
        "max_score": max_score,
        "is_correct": is_correct,
        "total_score": total_score,
        "regrade": True,
        "is_essay": is_essay,
        "model_key": model_key,
        "grading_method": grading_method,
        "comment": comment,
        "grading": grading,
        "error_cause": error_cause,
        "knowledge_points": knowledge_points,
        "writing_attribution": writing_attribution,
        "graded_at": submission.graded_at.isoformat() if submission.graded_at else None,
    }



@router.get("/submissions/{submission_id}/questions/{question_id}/regrade-history")
def list_regrade_history(
    submission_id: int,
    question_id: int,
    detail: bool = False,
    limit: int = 100,
    offset: int = 0,
    current_user: RegisteredPerson = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取某题的重批改历史记录（按 created_at DESC 排序）

    权限：仅教师/管理员（且必须是该 submission 所属考试的教师，admin 可读任意）
    入参：
    - detail: true 时返回完整 grading_json/writing_attribution_json/student_text
    - limit: 最多返回条数（默认 100，上限 500）
    - offset: 分页偏移

    返回：{submission_id, question_id, total, limit, offset, records: [...]}
    """
    from backend.models.tables import ExamSubmission, AnswerRegradeHistory

    if current_user.role not in ("teacher", "admin"):
        raise HTTPException(403, "仅教师/管理员可调用此接口")

    # 1. 校验 submission
    submission = db.query(ExamSubmission).filter(ExamSubmission.id == submission_id).first()
    if not submission:
        raise HTTPException(404, f"提交 {submission_id} 不存在")

    # 2. 校验当前用户是该 submission 所属考试的教师
    exam = db.query(Exam).filter(Exam.id == submission.exam_id).first()
    if not exam:
        raise HTTPException(404, "提交关联的考试不存在")
    if exam.teacher_id != current_user.id and current_user.role != "admin":
        raise HTTPException(403, "无权查看此提交的历史（仅该考试的教师可查看）")

    # 3. 校验 question 存在 + 属于该考试
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(404, f"题目 {question_id} 不存在")
    if question.exam_id != submission.exam_id:
        raise HTTPException(400, "题目不属于该提交对应的考试")

    # 4. 查询历史记录（按 created_at DESC 排序）
    query = db.query(AnswerRegradeHistory).filter(
        AnswerRegradeHistory.submission_id == submission_id,
        AnswerRegradeHistory.question_id == question_id,
    ).order_by(AnswerRegradeHistory.created_at.desc())

    total = query.count()
    safe_limit = max(1, min(int(limit), 500))
    safe_offset = max(0, int(offset))
    records = query.offset(safe_offset).limit(safe_limit).all()

    return {
        "submission_id": submission_id,
        "question_id": question_id,
        "total": total,
        "limit": safe_limit,
        "offset": safe_offset,
        "records": [_serialize_regrade_history(r, detail=detail) for r in records],
    }



def _serialize_regrade_history(r, detail: bool = False) -> dict:
    """序列化重批改历史记录

    detail=False（列表模式）：省略 grading_json/writing_attribution_json/完整 student_text，
                              仅返回 student_text_head 前 100 字 + knowledge_points 反序列化
    detail=True（详情模式）：返回完整字段
    """
    item = {
        "id": r.id,
        "operator_id": r.operator_id,
        "operator_name": r.operator.name if r.operator else None,
        "operator_role": r.operator.role if r.operator else None,
        "regrade_method": r.regrade_method,
        "input_mode": r.input_mode,
        "force_essay": r.force_essay,
        "before_score": r.before_score,
        "after_score": r.after_score,
        "before_is_correct": r.before_is_correct,
        "after_is_correct": r.after_is_correct,
        "max_score": r.max_score,
        "before_total_score": r.before_total_score,
        "after_total_score": r.after_total_score,
        "is_essay": r.is_essay,
        "model_key": r.model_key,
        "grading_method": r.grading_method,
        "error_cause": r.error_cause,
        "comment": r.comment,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }

    # 列表模式：student_text 截前 100 字，knowledge_points 反序列化
    item["student_text_head"] = (r.student_text or "")[:100]
    if r.knowledge_points_json:
        try:
            item["knowledge_points"] = json.loads(r.knowledge_points_json)
        except Exception:
            item["knowledge_points"] = []
    else:
        item["knowledge_points"] = []

    if detail:
        # 详情模式：返回完整字段
        item["student_text"] = r.student_text
        item["grading_json"] = json.loads(r.grading_json) if r.grading_json else None
        item["writing_attribution_json"] = (
            json.loads(r.writing_attribution_json) if r.writing_attribution_json else None
        )
        item["knowledge_points_json_raw"] = r.knowledge_points_json

    return item



def _init_excel_styles():
    """延迟初始化 Excel 样式（避免 import 时失败）"""
    global _EXCEL_HEADER_FILL
    if _EXCEL_HEADER_FILL is not None:
        return
    from openpyxl.styles import PatternFill, Font, Alignment
    _EXCEL_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")



def _build_excel_workbook(db: Session, submission_id: int):
    """生成单个 submission 的 Excel Workbook（D 方案抽取的辅助函数）

    被 export_excel_report（单导出）和 export_excel_batch（批量 ZIP）共享复用。

    Args:
        db: 数据库会话
        submission_id: 提交 ID

    Returns:
        (workbook, student_name, exam_title, filename)

    Raises:
        HTTPException: 404 submission 不存在
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from backend.models.tables import ExamSubmission, Exam, Answer, RegisteredPerson, Question, KnowledgeAnalysis

    # 查询 submission
    submission = db.query(ExamSubmission).filter(ExamSubmission.id == submission_id).first()
    if not submission:
        raise HTTPException(404, f"提交 {submission_id} 不存在")

    # 查询考试和学生信息
    exam = db.query(Exam).filter(Exam.id == submission.exam_id).first()
    student = db.query(RegisteredPerson).filter(RegisteredPerson.id == submission.student_id).first()
    student_name = student.name if student else f"用户#{submission.student_id}"
    exam_title = exam.title if exam else f"考试#{submission.exam_id}"

    # 查询所有答案（按题目顺序）
    answers_query = (
        db.query(Answer, Question)
        .join(Question, Answer.question_id == Question.id)
        .filter(Answer.submission_id == submission_id)
        .order_by(Question.order)
        .all()
    )

    # 类型中文化
    type_map = {"single": "单选题", "multi": "多选题", "judge": "判断题", "fill": "填空题", "essay": "解答题/作文"}
    correct_map = {True: "✓ 正确", False: "✗ 错误", None: "— 未批改"}

    # 创建 Excel
    _init_excel_styles()
    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # === Sheet 1: 题目明细 ===
    ws1 = wb.active
    ws1.title = "题目明细"

    # 顶部信息
    ws1["A1"] = "学生姓名"
    ws1["B1"] = student_name
    ws1["C1"] = "考试名称"
    ws1["D1"] = exam_title
    ws1["E1"] = "总分"
    ws1["F1"] = f"{submission.score or 0}/{exam.total_score if exam else 100}"
    for col in "ABCDEF":
        ws1[f"{col}1"].font = Font(bold=True)

    # 表头（第3行）
    headers1 = ["题号", "题型", "题目内容", "学生答案", "标准答案", "得分", "满分", "批改状态", "错因", "评语"]
    for col_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 数据行
    for i, (ans, q) in enumerate(answers_query, start=4):
        # 学生答案：如果是选项索引，转为 A/B/C/D
        student_ans = ans.content or ""
        if q.type in ("single", "multi") and student_ans:
            try:
                if q.type == "single":
                    student_ans = chr(ord("A") + int(student_ans))
                else:
                    student_ans = "".join(chr(ord("A") + int(x)) for x in student_ans.split(","))
            except (ValueError, IndexError):
                pass  # 保留原始内容
        elif q.type == "judge" and student_ans:
            student_ans = "正确" if student_ans.lower() == "true" else "错误"

        # 标准答案同样转换
        std_ans = q.answer or ""
        if q.type in ("single", "multi") and std_ans:
            try:
                if q.type == "single":
                    std_ans = chr(ord("A") + int(std_ans))
                else:
                    std_ans = "".join(chr(ord("A") + int(x)) for x in std_ans.split(","))
            except (ValueError, IndexError):
                pass
        elif q.type == "judge" and std_ans:
            std_ans = "正确" if std_ans.lower() == "true" else "错误"

        row_data = [
            q.order,
            type_map.get(q.type, q.type),
            q.content[:200],
            student_ans,
            std_ans,
            ans.score or 0,
            q.score,
            correct_map.get(ans.is_correct, "— 未批改"),
            "",  # 错因（Answer 表没有，留空）
            "",  # 评语（Answer 表没有，留空）
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=i, column=col_idx, value=val)
            cell.alignment = left_align if col_idx in (3, 4, 5, 9, 10) else center_align
            if ans.is_correct is False and col_idx == 8:
                cell.fill = wrong_fill

    # 列宽
    col_widths = [6, 10, 50, 20, 20, 8, 8, 12, 15, 40]
    for idx, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = w

    # === Sheet 2: 错题汇总 ===
    ws2 = wb.create_sheet("错题汇总")
    ws2["A1"] = f"{student_name} 的错题汇总"
    ws2["A1"].font = Font(bold=True, size=14)
    headers2 = ["题号", "题型", "题目内容", "学生答案", "标准答案", "错因", "改进建议"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    row_idx = 4
    for ans, q in answers_query:
        if ans.is_correct is not False:
            continue
        # 学生答案转换
        student_ans = ans.content or ""
        if q.type in ("single", "multi") and student_ans:
            try:
                if q.type == "single":
                    student_ans = chr(ord("A") + int(student_ans))
                else:
                    student_ans = "".join(chr(ord("A") + int(x)) for x in student_ans.split(","))
            except (ValueError, IndexError):
                pass
        elif q.type == "judge" and student_ans:
            student_ans = "正确" if student_ans.lower() == "true" else "错误"
        std_ans = q.answer or ""

        row_data = [q.order, type_map.get(q.type, q.type), q.content[:200], student_ans, std_ans, "（详见归因报告）", "（详见归因报告）"]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = left_align if col_idx in (3, 4, 5, 6, 7) else center_align
        row_idx += 1

    if row_idx == 4:
        ws2.cell(row=4, column=1, value="（无错题）")

    col_widths2 = [6, 10, 50, 20, 20, 20, 40]
    for idx, w in enumerate(col_widths2, start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # === Sheet 3: 归因报告 ===
    ws3 = wb.create_sheet("归因报告")
    ws3["A1"] = f"{student_name} 的学情归因报告"
    ws3["A1"].font = Font(bold=True, size=14)

    analyses = (
        db.query(KnowledgeAnalysis)
        .filter(KnowledgeAnalysis.student_id == submission.student_id)
        .order_by(KnowledgeAnalysis.created_at.desc())
        .all()
    )

    if not analyses:
        ws3.cell(row=3, column=1, value="暂无归因分析数据")
    else:
        row_idx = 3
        for ka in analyses:
            # 分析类型标题
            type_label = "数学学情" if ka.analysis_type == "math" else "写作学情"
            ws3.cell(row=row_idx, column=1, value=f"【{type_label}】（生成时间：{ka.created_at.strftime('%Y-%m-%d %H:%M')}）")
            ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            row_idx += 1

            # 雷达图数据
            try:
                radar = json.loads(ka.radar_json) if ka.radar_json else {}
            except json.JSONDecodeError:
                radar = {}
            if radar:
                ws3.cell(row=row_idx, column=1, value="维度掌握度：")
                ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                row_idx += 1
                for dim_name, score in radar.items():
                    ws3.cell(row=row_idx, column=1, value=dim_name)
                    ws3.cell(row=row_idx, column=2, value=f"{score*100:.0f}%")
                    row_idx += 1
                row_idx += 1

            # 薄弱点
            try:
                weak_points = json.loads(ka.weak_points_json) if ka.weak_points_json else {}
            except json.JSONDecodeError:
                weak_points = {}

            if ka.analysis_type == "math":
                # 数学：weak_points 是列表
                if isinstance(weak_points, list) and weak_points:
                    ws3.cell(row=row_idx, column=1, value="薄弱知识点：")
                    ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                    row_idx += 1
                    # 表头
                    for col_idx, h in enumerate(["知识点", "薄弱度", "错题数", "改进建议"], start=1):
                        cell = ws3.cell(row=row_idx, column=col_idx, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                    row_idx += 1
                    for wp in weak_points:
                        ws3.cell(row=row_idx, column=1, value=wp.get("knowledge_name", ""))
                        ws3.cell(row=row_idx, column=2, value=f"{wp.get('weakness_score', 0)*100:.0f}%")
                        ws3.cell(row=row_idx, column=3, value=wp.get("error_count", 0))
                        ws3.cell(row=row_idx, column=4, value=wp.get("suggestion", ""))
                        row_idx += 1
            else:
                # 写作：weak_points 是 dict（含 weak_dimensions/error_cause_distribution/overall_suggestion）
                if isinstance(weak_points, dict):
                    weak_dims = weak_points.get("weak_dimensions", [])
                    if weak_dims:
                        ws3.cell(row=row_idx, column=1, value="薄弱维度：")
                        ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                        row_idx += 1
                        for col_idx, h in enumerate(["维度", "薄弱度", "错因", "改进建议"], start=1):
                            cell = ws3.cell(row=row_idx, column=col_idx, value=h)
                            cell.fill = header_fill
                            cell.font = header_font
                        row_idx += 1
                        for wd in weak_dims:
                            ws3.cell(row=row_idx, column=1, value=wd.get("dimension_name", ""))
                            ws3.cell(row=row_idx, column=2, value=f"{wd.get('weakness_score', 0)*100:.0f}%")
                            ws3.cell(row=row_idx, column=3, value=", ".join(wd.get("error_causes", [])))
                            ws3.cell(row=row_idx, column=4, value=wd.get("suggestion", ""))
                            row_idx += 1
                    overall = weak_points.get("overall_suggestion", "")
                    if overall:
                        row_idx += 1
                        ws3.cell(row=row_idx, column=1, value="综合建议：")
                        ws3.cell(row=row_idx, column=1).font = Font(bold=True)
                        row_idx += 1
                        ws3.cell(row=row_idx, column=1, value=overall)
                        ws3.cell(row=row_idx, column=1).alignment = left_align
                        ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                        row_idx += 1

            row_idx += 2  # 不同分析类型之间空一行

    col_widths3 = [25, 15, 25, 50]
    for idx, w in enumerate(col_widths3, start=1):
        ws3.column_dimensions[get_column_letter(idx)].width = w

    # 文件名（中文安全）
    safe_student = student_name.replace(" ", "_").replace("/", "_")
    filename = f"答题卡报告_{safe_student}_{submission_id}.xlsx"

    return wb, student_name, exam_title, filename



@router.get("/export/excel/{submission_id}")
def export_excel_report(
    submission_id: int,
    current_user: RegisteredPerson = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """将整卷扫描批改结果导出为 Excel 报告

    包含 3 个 Sheet：
    1. 题目明细：题号、题型、题目内容、学生答案、标准答案、得分、是否正确、错因、评语
    2. 错题汇总：错题题号、题型、错因、改进建议
    3. 归因报告：薄弱知识点/写作薄弱维度 + 雷达图数据
    """
    import io
    from fastapi.responses import StreamingResponse

    if current_user.role not in ("teacher", "admin"):
        raise HTTPException(403, "仅教师/管理员可导出报告")

    # 校验提交归属（通过 submission → exam → teacher）
    from backend.models.tables import ExamSubmission
    submission = db.query(ExamSubmission).filter(ExamSubmission.id == submission_id).first()
    if not submission:
        raise HTTPException(404, f"提交 {submission_id} 不存在")
    exam = db.query(Exam).filter(Exam.id == submission.exam_id).first()
    if exam:
        assert_owner_or_admin(exam.teacher_id, current_user)

    wb, student_name, exam_title, filename = _build_excel_workbook(db, submission_id)

    # 保存到内存
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )



@router.get("/export/excel-batch")
def export_excel_batch(
    submission_ids: str,
    current_user: RegisteredPerson = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """D. 批量 Excel 报告导出（ZIP 包）

    一次导出多个 submission 的 Excel 报告，打包为 ZIP。

    Args:
        submission_ids: 逗号分隔的 submission ID 列表（最多 100 个）

    Returns:
        application/zip 流，包含 N 个 .xlsx 文件
    """
    import io
    import zipfile
    from fastapi.responses import StreamingResponse
    from backend.models.tables import ExamSubmission

    if current_user.role not in ("teacher", "admin"):
        raise HTTPException(403, "仅教师/管理员可批量导出报告")

    # 解析 submission_ids
    try:
        sid_list = [int(s.strip()) for s in submission_ids.split(",") if s.strip()]
    except ValueError:
        raise HTTPException(400, "submission_ids 格式错误，需为逗号分隔的整数列表")

    if not sid_list:
        raise HTTPException(400, "submission_ids 不能为空")
    if len(sid_list) > 100:
        raise HTTPException(400, f"单批最多导出 100 份报告，当前 {len(sid_list)} 份")

    # 校验所有 submission 存在 + 归属当前教师
    submissions = db.query(ExamSubmission).filter(ExamSubmission.id.in_(sid_list)).all()
    sub_map = {s.id: s for s in submissions}
    missing = [sid for sid in sid_list if sid not in sub_map]
    if missing:
        raise HTTPException(404, f"以下 submission 不存在: {missing[:10]}{'...' if len(missing) > 10 else ''}")

    # 校验每个 submission 的考试归属
    exam_ids = list(set(s.exam_id for s in submissions))
    exam_map = {e.id: e for e in db.query(Exam).filter(Exam.id.in_(exam_ids)).all()} if exam_ids else {}
    for sub in submissions:
        exam = exam_map.get(sub.exam_id)
        if exam:
            assert_owner_or_admin(exam.teacher_id, current_user)

    # 逐个生成 Excel 并打包成 ZIP
    zip_buf = io.BytesIO()
    used_filenames: set[str] = set()  # 处理同名冲突
    success_count = 0
    failed_list: list[dict] = []

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for sid in sid_list:
            try:
                wb, student_name, exam_title, original_filename = _build_excel_workbook(db, sid)
                # 处理同名冲突：如有重名，附加 _2/_3
                filename = original_filename
                if filename in used_filenames:
                    base, ext = os.path.splitext(original_filename)
                    n = 2
                    while f"{base}_{n}{ext}" in used_filenames:
                        n += 1
                    filename = f"{base}_{n}{ext}"
                used_filenames.add(filename)

                excel_buf = io.BytesIO()
                wb.save(excel_buf)
                excel_buf.seek(0)
                zf.writestr(filename, excel_buf.read())
                success_count += 1
            except HTTPException as e:
                failed_list.append({"submission_id": sid, "error": e.detail, "status_code": e.status_code})
            except Exception as e:
                failed_list.append({"submission_id": sid, "error": f"{type(e).__name__}: {e}"})
                logger.exception(
                    f"[AnswerSheetRoute] 批量导出 submission={sid} 异常: {e}"
                )

    zip_buf.seek(0)

    # 如果全部失败，返回错误
    if success_count == 0:
        raise HTTPException(
            500,
            f"所有 {len(sid_list)} 份报告生成失败: {failed_list[:3]}"
        )

    # ZIP 文件名（含时间戳）
    from datetime import datetime as _dt
    zip_filename = f"答题卡报告_batch_{_dt.now().strftime('%Y%m%d_%H%M%S')}.zip"

    logger.info(
        f"[AnswerSheetRoute] 批量导出完成: total={len(sid_list)}, "
        f"success={success_count}, failed={len(failed_list)}"
    )

    # 注意：失败的 submission 信息记录在 response header 中（避免破坏 ZIP 流）
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{zip_filename}",
        "X-Batch-Total": str(len(sid_list)),
        "X-Batch-Success": str(success_count),
        "X-Batch-Failed": str(len(failed_list)),
    }
    if failed_list:
        # 把失败列表序列化到 header（截断防止 header 过大）
        failed_summary = json.dumps(failed_list[:5], ensure_ascii=False)[:500]
        headers["X-Batch-Failed-Detail"] = failed_summary

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers=headers,
    )

